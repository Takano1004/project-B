from flask import render_template, jsonify
from sqlalchemy import func
from model import db, ユーザ, 所属, 拾得物, 拾得物分類, 拾得物管理状況
from datetime import datetime
import os
import openpyxl

def top():
    return render_template("top.html")

def json(data):
    if data == "user":
        j = j_user()
    elif data == "item":
        j = j_item()
    elif data == "category":
        j = j_category()
    elif data == "dept":
        j = j_dept()
    return(jsonify(j))

def j_dept():
    data = db.session.query(所属).all()
    obj = []
    for i in data:
        obj.append({
            "所属ID":i.ID,
            "所属分類":i.所属分類,
            "所属名":i.所属名, 
            })
    return(obj)

def j_user():
    data = db.session.query(ユーザ,所属).filter(ユーザ.所属ID==所属.ID).all()
    obj = []
    for i,j in data:
        obj.append({
            "ID":i.ID,
            "氏名":i.氏名,
            "電話番号":i.電話番号,
            "メールアドレス":i.メールアドレス,
            "住所":i.住所,
            "所属分類":j.所属分類,
            "所属名":j.所属名, 
            })
    return(obj)

def j_item(key=""):
    lastdt = db.session.query(拾得物管理状況.拾得物ID.label("oid"),func.max(拾得物管理状況.変更日時).label("last")).group_by(拾得物管理状況.拾得物ID).subquery()
    tmp = db.session.query(拾得物管理状況,拾得物,拾得物分類,ユーザ,所属).filter(拾得物管理状況.拾得物ID==拾得物.ID).filter(拾得物.拾得物分類ID==拾得物分類.ID).filter(拾得物管理状況.ユーザID==ユーザ.ID).filter(ユーザ.所属ID==所属.ID).filter(拾得物管理状況.変更日時==lastdt.c.last).filter(拾得物管理状況.拾得物ID==lastdt.c.oid)
    if key == "":
        data = tmp.all()
    else:
        data = tmp.filter(拾得物分類.物品名.contains(key)).all()
    obj = []
    for i,j,k,l,m in data:
        obj.append({
            "大分類":k.大分類,
            "物品名":k.物品名,
            "貴重品":k.貴重品,
            "拾得物ID":j.ID,
            "拾得場所":j.拾得場所,
            "色":j.色,
            "特徴":j.特徴,
            "画像":j.画像,
            "拾得物管理状況ID":i.ID,
            "変更日時":i.変更日時,
            "変更内容":i.変更内容,
            "氏名":l.氏名,
            "所属名":m.所属名, 
            })
    return(obj)

def j_category():
    data = db.session.query(拾得物分類).all()
    obj = []
    for i in data:
        obj.append({
            "ID":i.ID,
            "大分類":i.大分類,
            "物品名":i.物品名,
            "頭1":i.頭1,
            "頭2":i.頭2,
            "五十音":i.五十音,
            "貴重品":i.貴重品, 
            })
    return(obj)

def f_item(k):
    j = j_item(k)
    return render_template("item_list.html", j=j, n=len(j))

def d_item(id):
    data = db.session.query(拾得物管理状況,ユーザ).filter(拾得物管理状況.ユーザID==ユーザ.ID).filter(拾得物管理状況.拾得物ID==id)
    list = []
    for i,j in data:
        list.append({
            "変更日時":i.変更日時,
            "変更内容":i.変更内容,
            "氏名":j.氏名,
        })
    i,j = db.session.query(拾得物,拾得物分類).filter(拾得物.拾得物分類ID==拾得物分類.ID).filter(拾得物.ID==id).first()
    if i.画像 is None:
        i.画像 = ""
    d = {
        "大分類":j.大分類,
        "物品名":j.物品名,
        "拾得場所":i.拾得場所,
        "色":i.色,
        "特徴":i.特徴,
        "画像":i.画像, 
        "変更履歴":list, 
        "id":id
        }
    return render_template("item_detail.html", d=d, u=j_user())

def user_form():
    return render_template("reg_user.html", j=j_dept())

def user_reg(d):
    tmp = ユーザ(
        氏名 = d["氏名"],
        所属ID = d["所属ID"],
        電話番号 = d["電話番号"],
        メールアドレス = d["メールアドレス"],
        住所= "",
    )
    db.session.add(tmp)
    db.session.commit()

def item_form():
    return render_template("reg_item.html", j=j_category(), j2=j_user())

def item_reg(d, f):
    tmp = 拾得物(
        拾得物分類ID = d["拾得物分類ID"],
        拾得場所 = d["拾得場所"],
        色 = d["色"],
        特徴 = d["特徴"],
        画像 = "",
    )
    db.session.add(tmp)
    db.session.flush()
    n,ext = os.path.splitext(f.filename)
    fn = f"{tmp.ID}{ext}"
    tmp.画像 = fn
    f.save(os.path.join("./static/img/", fn))
    tmp2 = 拾得物管理状況(
        ユーザID = d["ユーザID"],
        拾得物ID = tmp.ID,
        変更日時 = datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
        変更内容 = "拾得",
    )
    db.session.add(tmp2)
    db.session.commit()

def act_reg(d):
    tmp = 拾得物管理状況(
        ユーザID = d["ユーザID"],
        拾得物ID = d["拾得物ID"],
        変更日時 = datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
        変更内容 = d["変更内容"],
    )
    db.session.add(tmp)
    db.session.commit()

def req_item():
    return render_template("req_item.html")

def req_list():
    return render_template("req_list.html")

def dl():
    return render_template("dl.html")

def dl_data():
    lastdt = db.session.query(拾得物管理状況.拾得物ID.label("oid"),func.max(拾得物管理状況.変更日時).label("last")).group_by(拾得物管理状況.拾得物ID).subquery()
    tmp = db.session.query(拾得物管理状況,拾得物,拾得物分類).filter(拾得物管理状況.拾得物ID==拾得物.ID).filter(拾得物.拾得物分類ID==拾得物分類.ID).filter(拾得物管理状況.変更日時==lastdt.c.last).filter(拾得物管理状況.拾得物ID==lastdt.c.oid).filter(拾得物管理状況.変更内容=='拾得').all()
    obj = []
    for i,j,k in tmp:
        obj.append({
            "物品名":k.物品名,
            "拾得物ID":j.ID,
            "拾得場所":j.拾得場所,
            "色":j.色,
            "特徴":j.特徴,
            "拾得物管理状況ID":i.ID,
            "変更日時":i.変更日時,
            "変更内容":i.変更内容,
            })
    path = os.path.dirname(__file__)
    wb = openpyxl.load_workbook(f"{path}\style05.xlsx")
    pat = ["0","1","2","3","4","5","6","7","8","9","０","１","２","３","４","５","６","７","８","９"]
    i,p = 0,0
    for row in obj:
        if (i + 10) > 46 or i == 0:
            i,p = 0,p+1
            ws = wb.copy_worksheet(wb["template"])
            ws.title = f"拾得物リスト{p}"
        ws.cell(10+i,9).value = row["物品名"]
        if row["物品名"] == "現金":
            j = 0
            for s in row["特徴"][::-1]:
                if not(s in pat): continue
                ws.cell(10+i,8-j).value = s
                print(s)
                j += 1
            ws.cell(10+i,8-j).value = "¥"
        else:
            ws.cell(10+i,11).value = row["特徴"] + row["色"]
        ws.cell(10+i,17).value = row["変更日時"].strftime('%m/%d %H:%M')
        ws.cell(10+i,19).value = row["拾得場所"]
        i += 4
    ws = wb.remove(wb["template"])
    wb.save(f"{path}\out.xlsx")
    return